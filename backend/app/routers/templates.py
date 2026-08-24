import json
import uuid
from datetime import datetime, timezone
from pathlib import Path, PurePosixPath, PureWindowsPath

from fastapi import APIRouter, HTTPException, UploadFile
from openpyxl import load_workbook

from app.config import settings
from app.models.schemas import TemplateInfo

router = APIRouter(prefix="/api/templates", tags=["templates"])


def _meta_path(template_id: str) -> Path:
    return settings.templates_dir / f"{template_id}_meta.json"


def _resolve_template_path(meta: dict) -> Path | None:
    """Resolve current and legacy template metadata on every supported OS.

    Older metadata stored an absolute path.  Keeping only its final filename
    makes a library copied between macOS, Linux, and Windows portable and also
    prevents metadata from escaping the configured template directory.
    """
    stored_path = str(meta.get("path", "")).strip()
    if not stored_path:
        return None
    filename = PureWindowsPath(PurePosixPath(stored_path).name).name
    if not filename or filename in {".", ".."}:
        return None
    return settings.templates_dir / filename


def _all_templates() -> list[dict]:
    templates = []
    for p in sorted(settings.templates_dir.glob("*_meta.json")):
        meta = json.loads(p.read_text(encoding="utf-8"))
        template_path = _resolve_template_path(meta)
        if template_path is not None and meta.get("path") != template_path.name:
            meta["path"] = template_path.name
            p.write_text(json.dumps(meta, ensure_ascii=False), encoding="utf-8")
        if "available_sheets" not in meta and template_path is not None and template_path.is_file():
            meta["available_sheets"] = _detect_available_sheets(template_path)
            p.write_text(json.dumps(meta, ensure_ascii=False), encoding="utf-8")
        templates.append(meta)
    return templates


def _detect_available_sheets(template_path: Path) -> list[int]:
    implemented = {
        2: "2.",
        3: "3.",
        6: "6.",
        7: "7.",
        8: "8.",
        9: "9.",
    }
    wb = load_workbook(template_path, read_only=True)
    sheet_names = [name.strip() for name in wb.sheetnames]
    wb.close()
    return [
        sheet_number
        for sheet_number, prefix in implemented.items()
        if any(name.startswith(prefix) for name in sheet_names)
    ]


@router.get("", response_model=dict)
def list_templates():
    return {"templates": [TemplateInfo(**template).model_dump() for template in _all_templates()]}


@router.post("/upload", response_model=TemplateInfo, status_code=201)
async def upload_template(file: UploadFile):
    if not file.filename.lower().endswith(".xlsx"):
        raise HTTPException(status_code=400, detail="Only .xlsx files are accepted")

    tid = str(uuid.uuid4())
    dest = settings.templates_dir / f"{tid}.xlsx"
    dest.write_bytes(await file.read())

    meta = {
        "template_id": tid,
        "name": file.filename,
        "path": dest.name,
        "available_sheets": _detect_available_sheets(dest),
        "uploaded_at": datetime.now(timezone.utc).isoformat(),
    }
    _meta_path(tid).write_text(json.dumps(meta, ensure_ascii=False), encoding="utf-8")
    return TemplateInfo(**meta)


@router.delete("/{template_id}", status_code=204)
def delete_template(template_id: str):
    meta_path = _meta_path(template_id)
    if not meta_path.exists():
        return

    meta = json.loads(meta_path.read_text(encoding="utf-8"))
    template_path = _resolve_template_path(meta)

    if template_path is not None and template_path.is_file():
        template_path.unlink()

    meta_path.unlink()


def get_template_path(template_id: str) -> Path:
    mp = _meta_path(template_id)
    if not mp.exists():
        raise HTTPException(status_code=404, detail="Template not found")
    meta = json.loads(mp.read_text(encoding="utf-8"))
    p = _resolve_template_path(meta)
    if p is None or not p.is_file():
        raise HTTPException(status_code=404, detail="Template file missing on disk")
    return p
