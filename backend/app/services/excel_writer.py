"""
Excel writer.

Copies the Form 107-A template and fills in all target sheets while
preserving formatting, styles, and data validations.

The Sheet 7 "no subservice" checkbox is a VML Form Control that openpyxl
cannot touch, so it is toggled via direct ZIP/XML manipulation.
"""
import shutil
import zipfile
from pathlib import Path
import re
from copy import copy
from xml.etree import ElementTree as ET

import openpyxl
from openpyxl.cell.cell import MergedCell

from app.models.extraction import ExtractedFormData, ITGCSection


def _cell(ws, row: int, col: int):
    """Return a writable cell, resolving merged ranges to their top-left cell."""
    c = ws.cell(row, col)
    if isinstance(c, MergedCell):
        for rng in ws.merged_cells.ranges:
            if rng.min_row <= row <= rng.max_row and rng.min_col <= col <= rng.max_col:
                return ws.cell(rng.min_row, rng.min_col)
    return c


# ── Sheet 6 row map ───────────────────────────────────────────────────────────

_S6 = {
    "change_mgmt":    {"has_desc": 6,  "page_refs": 7,  "sec_b": 10, "sec_c": 12, "risks": [13, 15]},
    "access_mgmt":    {"has_desc": 19, "page_refs": 20, "sec_b": 23, "sec_c": 25, "risks": [26, 28, 30]},
    "job_scheduling": {"has_desc": 34, "page_refs": 35, "sec_b": 38, "sec_c": 40, "risks": [41, 43, 45]},
}
_COL_D = 4  # (X)  dropdown "See col (XI)" / "None"
_COL_E = 5  # (XI) free-text reply / control IDs


# ── Sheet writers ─────────────────────────────────────────────────────────────

def _write_sheet2(ws, data: ExtractedFormData) -> None:
    ws["C3"] = data.sheet2.report_name
    ws["D5"] = data.sheet2.period


def _write_sheet3(ws, data: ExtractedFormData) -> None:
    s3 = data.sheet3
    if s3.has_qualified_opinion and s3.opinion:
        _cell(ws, 4, 1).value = s3.opinion.description
        _cell(ws, 4, 4).value = s3.opinion.explanation
    for i, exc in enumerate(s3.exceptions):
        r = 12 + i
        _cell(ws, r, 1).value = exc.index
        _cell(ws, r, 2).value = exc.control
        _cell(ws, r, 3).value = exc.exception_desc
        _cell(ws, r, 4).value = exc.mgmt_response
        _cell(ws, r, 5).value = exc.is_audited
        _cell(ws, r, 6).value = exc.audit_relevant


def _write_itgc_section(ws, section: ITGCSection, rows: dict) -> None:
    def _set_if_empty(row: int, col: int, value) -> None:
        c = _cell(ws, row, col)
        if not c.value:
            c.value = value

    # Preserve all pre-filled template values; only fill blank cells
    _set_if_empty(rows["has_desc"], _COL_E, section.has_process_description)
    _set_if_empty(rows["sec_b"],    _COL_E, section.section_b_applicable)
    _set_if_empty(rows["sec_c"],    _COL_E, section.section_c_applicable)

    # Page refs: write when either LLM or template says "Yes"
    effective_has_desc = _cell(ws, rows["has_desc"], _COL_E).value or section.has_process_description
    if effective_has_desc == "Yes" and section.page_refs:
        _set_if_empty(rows["page_refs"], _COL_E, section.page_refs)

    # Control IDs: write whenever section_c is Applicable (template always pre-fills this)
    effective_sec_c = _cell(ws, rows["sec_c"], _COL_E).value or section.section_c_applicable
    if effective_sec_c == "Applicable":
        for risk_row, ids in zip(rows["risks"], section.risk_control_ids):
            _set_if_empty(risk_row, _COL_D, "See col (XI)")
            if ids:
                _cell(ws, risk_row, _COL_E).value = ids


def _write_sheet6(ws, data: ExtractedFormData) -> None:
    _write_itgc_section(ws, data.sheet6.change_mgmt,   _S6["change_mgmt"])
    _write_itgc_section(ws, data.sheet6.access_mgmt,   _S6["access_mgmt"])
    _write_itgc_section(ws, data.sheet6.job_scheduling, _S6["job_scheduling"])


def _write_sheet7(ws, data: ExtractedFormData) -> None:
    for i, org in enumerate(data.sheet7.organizations):
        _cell(ws, 4 + i, 1).value = org.name
        _cell(ws, 4 + i, 2).value = org.services


def _write_sheet8(ws, data: ExtractedFormData) -> None:
    for i, cuec in enumerate(data.sheet8.cuecs):
        _cell(ws, 5 + i, 1).value = cuec.objective_and_page
        _cell(ws, 5 + i, 2).value = cuec.description


def _copy_row_style(ws, source_row: int, target_row: int, max_col: int) -> None:
    ws.row_dimensions[target_row].height = ws.row_dimensions[source_row].height
    for col in range(1, max_col + 1):
        source = ws.cell(source_row, col)
        target = ws.cell(target_row, col)
        if source.has_style:
            target._style = copy(source._style)
        if source.number_format:
            target.number_format = source.number_format
        if source.alignment:
            target.alignment = copy(source.alignment)
        if source.protection:
            target.protection = copy(source.protection)


def _insert_rows_preserving_merges(ws, row: int, amount: int) -> None:
    shifted_ranges = []
    for merged_range in list(ws.merged_cells.ranges):
        if merged_range.min_row >= row:
            shifted_ranges.append((
                merged_range.min_row,
                merged_range.min_col,
                merged_range.max_row,
                merged_range.max_col,
            ))
            ws.unmerge_cells(str(merged_range))

    ws.insert_rows(row, amount)

    for min_row, min_col, max_row, max_col in shifted_ranges:
        ws.merge_cells(
            start_row=min_row + amount,
            start_column=min_col,
            end_row=max_row + amount,
            end_column=max_col,
        )


def _write_sheet9(ws, data: ExtractedFormData) -> None:
    csocs = data.sheet9.csocs
    start_row = 5
    reserved_rows = 5
    if len(csocs) > reserved_rows:
        rows_to_insert = len(csocs) - reserved_rows
        insert_at = start_row + reserved_rows
        _insert_rows_preserving_merges(ws, insert_at, rows_to_insert)
        for row in range(insert_at, insert_at + rows_to_insert):
            _copy_row_style(ws, start_row + reserved_rows - 1, row, 10)

    for i, csoc in enumerate(csocs):
        row = start_row + i
        _cell(ws, row, 1).value = csoc.objective_and_page
        _cell(ws, row, 2).value = csoc.subservice_org
        _cell(ws, row, 3).value = csoc.relevant
        _cell(ws, row, 4).value = csoc.description
        if csoc.necessary:
            _cell(ws, row, 5).value = csoc.necessary
        if csoc.reason:
            _cell(ws, row, 6).value = csoc.reason
        if csoc.response:
            _cell(ws, row, 7).value = csoc.response


# ── VML checkbox helpers ──────────────────────────────────────────────────────

_CONTROL_WORKSHEETS = {
    "xl/worksheets/sheet3.xml",
    "xl/worksheets/sheet7.xml",
    "xl/worksheets/sheet8.xml",
    "xl/worksheets/sheet9.xml",
}


def _rels_owner_worksheet(rel_path: str) -> str:
    return rel_path.replace("xl/worksheets/_rels/", "xl/worksheets/").replace(".rels", "")


def _worksheet_rels_path(worksheet_path: str) -> str:
    return worksheet_path.replace("xl/worksheets/", "xl/worksheets/_rels/") + ".rels"


def _resolve_rel_target(rel_path: str, target: str) -> str:
    source_dir = Path(rel_path).parent.parent
    parts: list[str] = []
    for part in (source_dir / target).as_posix().split("/"):
        if part == "..":
            if parts:
                parts.pop()
        elif part and part != ".":
            parts.append(part)
    return "/".join(parts)


def _control_related_parts(template_files: dict[str, bytes]) -> set[str]:
    parts = set(_CONTROL_WORKSHEETS)
    rel_ns = {"rel": "http://schemas.openxmlformats.org/package/2006/relationships"}

    for worksheet_path in _CONTROL_WORKSHEETS:
        rel_path = _worksheet_rels_path(worksheet_path)
        if rel_path not in template_files:
            continue
        parts.add(rel_path)
        root = ET.fromstring(template_files[rel_path])
        for rel in root.findall("rel:Relationship", rel_ns):
            target = rel.attrib.get("Target", "")
            if not target or "://" in target:
                continue
            resolved = _resolve_rel_target(rel_path, target)
            if resolved in template_files:
                parts.add(resolved)

    return parts


def _merge_worksheet_root_attrs(output_xml: str, template_xml: str) -> str:
    output_root_match = re.search(r"<worksheet\b[^>]*>", output_xml)
    template_root_match = re.search(r"<worksheet\b[^>]*>", template_xml)
    if not output_root_match or not template_root_match:
        return output_xml

    output_root = output_root_match.group(0)
    template_root = template_root_match.group(0)

    for attr in re.findall(r'\s(?:xmlns:[\w]+|mc:Ignorable|xr:uid)="[^"]*"', template_root):
        attr_name = attr.strip().split("=", 1)[0]
        output_root = re.sub(rf'\s{re.escape(attr_name)}="[^"]*"', "", output_root)
        output_root = output_root[:-1] + attr + ">"

    return (
        output_xml[:output_root_match.start()]
        + output_root
        + output_xml[output_root_match.end():]
    )


def _add_worksheet_control_namespaces(output_xml: str) -> str:
    root_match = re.search(r"<worksheet\b[^>]*>", output_xml)
    if not root_match:
        return output_xml

    root = root_match.group(0)
    additions = {
        "xmlns:r": 'xmlns:r="http://schemas.openxmlformats.org/officeDocument/2006/relationships"',
        "xmlns:xdr": 'xmlns:xdr="http://schemas.openxmlformats.org/drawingml/2006/spreadsheetDrawing"',
        "xmlns:x14": 'xmlns:x14="http://schemas.microsoft.com/office/spreadsheetml/2009/9/main"',
        "xmlns:mc": 'xmlns:mc="http://schemas.openxmlformats.org/markup-compatibility/2006"',
        "xmlns:x14ac": 'xmlns:x14ac="http://schemas.microsoft.com/office/spreadsheetml/2009/9/ac"',
        "xmlns:xr": 'xmlns:xr="http://schemas.microsoft.com/office/spreadsheetml/2014/revision"',
        "xmlns:xr2": 'xmlns:xr2="http://schemas.microsoft.com/office/spreadsheetml/2015/revision2"',
        "xmlns:xr3": 'xmlns:xr3="http://schemas.microsoft.com/office/spreadsheetml/2016/revision3"',
        "xmlns:xm": 'xmlns:xm="http://schemas.microsoft.com/office/excel/2006/main"',
        "mc:Ignorable": 'mc:Ignorable="x14ac xr xr2 xr3"',
    }

    for marker, attr in additions.items():
        if marker not in root:
            root = root[:-1] + f" {attr}>"

    root = re.sub(
        r'mc:Ignorable="[^"]*"',
        'mc:Ignorable="x14ac xr xr2 xr3"',
        root,
    )

    return output_xml[:root_match.start()] + root + output_xml[root_match.end():]


def _extract_worksheet_control_tail(template_xml: str) -> str:
    start = template_xml.find("<legacyDrawing")
    if start == -1:
        return ""
    end = template_xml.rfind("</worksheet>")
    if end == -1:
        return ""
    return template_xml[start:end]


def _restore_worksheet_controls(files: dict[str, bytes], template_files: dict[str, bytes]) -> None:
    for name in _CONTROL_WORKSHEETS:
        if name not in template_files:
            continue
        template_content = template_files[name]

        tail = _extract_worksheet_control_tail(template_content.decode("utf-8"))
        if not tail or name not in files:
            continue

        output_xml = files[name].decode("utf-8")
        if "<legacyDrawing" in output_xml:
            continue

        output_xml = _merge_worksheet_root_attrs(output_xml, template_content.decode("utf-8"))
        output_xml = output_xml.replace("</worksheet>", f"{tail}</worksheet>")
        files[name] = output_xml.encode("utf-8")


def _merge_content_types(
    files: dict[str, bytes],
    template_files: dict[str, bytes],
    copied_parts: set[str],
) -> None:
    if "[Content_Types].xml" not in files or "[Content_Types].xml" not in template_files:
        return

    output_xml = files["[Content_Types].xml"].decode("utf-8")
    template_xml = template_files["[Content_Types].xml"].decode("utf-8")

    needed_tags = []
    for tag in re.findall(r"<Default\b[^>]*/>", template_xml):
        if 'Extension="vml"' in tag or 'Extension="bin"' in tag:
            needed_tags.append(tag)
    for tag in re.findall(r"<Override\b[^>]*/>", template_xml):
        part_match = re.search(r'PartName="/([^"]+)"', tag)
        if part_match and part_match.group(1) in copied_parts:
            needed_tags.append(tag)

    for tag in needed_tags:
        part_match = re.search(r'(?:PartName|Extension)="([^"]+)"', tag)
        marker = part_match.group(1) if part_match else tag
        if marker not in output_xml:
            output_xml = output_xml.replace("</Types>", f"{tag}</Types>")

    files["[Content_Types].xml"] = output_xml.encode("utf-8")


def _set_checkbox_checked(xml: str, checked: bool) -> str:
    if checked:
        if 'checked="Checked"' in xml:
            return xml
        return xml.replace('objectType="CheckBox"', 'objectType="CheckBox" checked="Checked"')
    return re.sub(r'\schecked="[^"]+"', "", xml)


def _restore_excel_controls(
    template_path: Path,
    output_path: Path,
    *,
    tick_unqualified_opinion: bool,
    tick_no_test_exceptions: bool,
    tick_no_subservice: bool,
) -> None:
    with zipfile.ZipFile(template_path, "r") as zin:
        template_files = {name: zin.read(name) for name in zin.namelist()}
    with zipfile.ZipFile(output_path, "r") as zin:
        files = {name: zin.read(name) for name in zin.namelist()}

    copied_parts = _control_related_parts(template_files)
    for name in copied_parts:
        if name in template_files and name not in _CONTROL_WORKSHEETS:
            files[name] = template_files[name]

    _restore_worksheet_controls(files, template_files)
    _merge_content_types(files, template_files, copied_parts)

    checkbox_states = {
        # Sheet 3: no test exceptions/deviations
        "xl/ctrlProps/ctrlProp1.xml": tick_no_test_exceptions,
        # Sheet 3: unqualified opinion
        "xl/ctrlProps/ctrlProp2.xml": tick_unqualified_opinion,
        # Sheet 7: no indexed subservice organizations
        "xl/ctrlProps/ctrlProp8.xml": tick_no_subservice,
    }

    for target, checked in checkbox_states.items():
        if target in files:
            files[target] = _set_checkbox_checked(
                files[target].decode("utf-8"),
                checked,
            ).encode("utf-8")

    with zipfile.ZipFile(output_path, "w", zipfile.ZIP_DEFLATED) as zout:
        for name, content in files.items():
            zout.writestr(name, content)


# ── Main entry ────────────────────────────────────────────────────────────────

def write_excel(
    data: ExtractedFormData,
    output_path: Path,
    template_path: Path,
) -> Path:
    shutil.copy2(template_path, output_path)
    wb = openpyxl.load_workbook(output_path)

    ws = {name.strip(): wb[name] for name in wb.sheetnames}

    if data.sheet2:
        _write_sheet2(ws["2.索引信息"], data)
    if data.sheet3:
        _write_sheet3(ws["3.报告保留意见和测试异常情况"], data)
    if data.sheet6:
        _write_sheet6(ws["6.IT流程和IT一般控制"], data)
    if data.sheet7:
        _write_sheet7(ws["7.子服务机构"], data)
    if data.sheet8:
        _write_sheet8(ws["8.补偿性用户实体控制"], data)
    if data.sheet9:
        _write_sheet9(ws["9.补偿性分包服务机构控制"], data)

    wb.save(output_path)

    _restore_excel_controls(
        template_path,
        output_path,
        tick_unqualified_opinion=bool(
            data.sheet3 and not data.sheet3.has_qualified_opinion
        ),
        tick_no_test_exceptions=bool(
            data.sheet3 and not data.sheet3.exceptions
        ),
        tick_no_subservice=bool(data.sheet7 and not data.sheet7.has_subservice),
    )

    return output_path
